"""
PDF Corruption Checker (Recursive) — Production Ready
------------------------------------------------------
Scans all PDF files in a specified folder and all subfolders,
validates each file, prints a live table, and exports an Excel report.

Usage:
    python check_pdf_corruption.py <folder_path> [--out report.xlsx] [--workers N]
    python check_pdf_corruption.py               # prompts for folder path

Requirements:
    pip install pypdf openpyxl
"""

from __future__ import annotations

import argparse
import concurrent.futures
import logging
import os
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import List, Optional

# ── Dependency bootstrap ──────────────────────────────────────────────────────
def _ensure(*packages: str) -> None:
    import importlib, subprocess
    for pkg in packages:
        if importlib.util.find_spec(pkg.replace("-", "_")) is None:
            print(f"Installing required library: {pkg} …")
            subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])

_ensure("pypdf", "openpyxl")

from pypdf import PdfReader                                      # noqa: E402
from pypdf.errors import PdfReadError, PdfStreamError            # noqa: E402
import openpyxl                                                  # noqa: E402
from openpyxl.styles import (                                    # noqa: E402
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter                     # noqa: E402

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.WARNING,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

# ── ANSI colours (disabled on non-TTY / Windows without ANSI support) ─────────
def _supports_colour() -> bool:
    if os.environ.get("NO_COLOR"):
        return False
    return hasattr(sys.stdout, "isatty") and sys.stdout.isatty()

_CLR = _supports_colour()
GREEN  = "\033[92m" if _CLR else ""
RED    = "\033[91m" if _CLR else ""
YELLOW = "\033[93m" if _CLR else ""
CYAN   = "\033[96m" if _CLR else ""
BOLD   = "\033[1m"  if _CLR else ""
RESET  = "\033[0m"  if _CLR else ""


# ── PDF validation ────────────────────────────────────────────────────────────
STATUS_OK       = "OK"
STATUS_CORRUPT  = "CORRUPT"
STATUS_EMPTY    = "EMPTY"
STATUS_NOT_PDF  = "NOT_A_PDF"
STATUS_NO_PAGES = "NO_PAGES"

def check_pdf(filepath: Path, root: Optional[Path] = None) -> dict:
    """
    Validate a single PDF file.

    Returns a dict:
        filepath     : Path
        relative_path: str   (relative to root, or absolute if root is None)
        filename     : str
        folder       : str
        status       : "OK" | "CORRUPT" | "EMPTY" | "NOT_A_PDF" | "NO_PAGES"
        pages        : int   (0 if unreadable)
        size_kb      : float
        reason       : str   (empty string if OK)
        checked_at   : str   (ISO-8601 timestamp)
        duration_ms  : float (time taken in milliseconds)
    """
    t0 = time.perf_counter()
    checked_at = datetime.now().isoformat(timespec="seconds")

    try:
        stat = filepath.stat()
        size_kb = round(stat.st_size / 1024, 2)
    except OSError as exc:
        size_kb = 0.0
        logger.warning("Cannot stat %s: %s", filepath, exc)

    rel = str(filepath.relative_to(root)) if root else str(filepath)

    result: dict = {
        "filepath"     : filepath,
        "relative_path": rel,
        "filename"     : filepath.name,
        "folder"       : str(filepath.parent.relative_to(root)) if root else str(filepath.parent),
        "status"       : STATUS_OK,
        "pages"        : 0,
        "size_kb"      : size_kb,
        "reason"       : "",
        "checked_at"   : checked_at,
        "duration_ms"  : 0.0,
    }

    def _finish(status: str, reason: str = "") -> dict:
        result["status"]      = status
        result["reason"]      = reason
        result["duration_ms"] = round((time.perf_counter() - t0) * 1000, 1)
        return result

    # 1. Empty file
    if size_kb == 0:
        return _finish(STATUS_EMPTY, "File is 0 bytes")

    # 2. Magic-byte check (%PDF header)
    try:
        with open(filepath, "rb") as fh:
            header = fh.read(5)
        if not header.startswith(b"%PDF"):
            return _finish(STATUS_NOT_PDF, f"Invalid PDF header: {header[:8]!r}")
    except OSError as exc:
        return _finish(STATUS_CORRUPT, f"Cannot read file: {exc}")

    # 3. Full structural check via pypdf
    try:
        reader = PdfReader(str(filepath), strict=False)
        page_count = len(reader.pages)

        if page_count == 0:
            return _finish(STATUS_NO_PAGES, "PDF has 0 pages")

        for i, page in enumerate(reader.pages):
            try:
                _ = page.extract_text()
            except Exception as page_exc:
                result["pages"] = page_count
                return _finish(STATUS_CORRUPT, f"Page {i + 1} unreadable: {page_exc}")

        result["pages"] = page_count

    except PdfReadError as exc:
        return _finish(STATUS_CORRUPT, f"PdfReadError: {exc}")
    except PdfStreamError as exc:
        return _finish(STATUS_CORRUPT, f"PdfStreamError: {exc}")
    except Exception as exc:
        logger.debug("Unexpected error in %s", filepath, exc_info=True)
        return _finish(STATUS_CORRUPT, f"Unexpected error: {exc}")

    return _finish(STATUS_OK)


# ── Recursive folder scan ─────────────────────────────────────────────────────
def collect_pdfs(root: Path) -> List[Path]:
    """Return deduplicated, sorted list of all PDF files under *root*."""
    seen, unique = set(), []
    for pattern in ("*.pdf", "*.PDF"):
        for p in sorted(root.rglob(pattern)):
            key = p.resolve()
            if key not in seen:
                seen.add(key)
                unique.append(p)
    return unique


def scan_folder(root: Path, workers: int = 4) -> List[dict]:
    """Walk *root* recursively, validate every PDF, print live progress."""
    pdfs = collect_pdfs(root)
    total = len(pdfs)

    if total == 0:
        print(f"\n{YELLOW}No PDF files found under: {root}{RESET}\n")
        return []

    print(f"\n{BOLD}{CYAN}Scanning {total} PDF file(s) under:{RESET} {root}\n")
    col_w = 110
    print(f"{'#':<5} {'Status':<12} {'Pages':>6}  {'Size (KB)':>10}  {'ms':>6}  Path")
    print("-" * col_w)

    results: List[dict] = [None] * total  # type: ignore[list-item]

    def _worker(args):
        idx, filepath = args
        return idx, check_pdf(filepath, root)

    with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(_worker, (i, p)): i for i, p in enumerate(pdfs)}
        for fut in concurrent.futures.as_completed(futures):
            try:
                idx, res = fut.result()
            except Exception as exc:
                idx = futures[fut]
                res = {
                    "filepath": pdfs[idx], "relative_path": str(pdfs[idx]),
                    "filename": pdfs[idx].name, "folder": str(pdfs[idx].parent),
                    "status": STATUS_CORRUPT, "pages": 0,
                    "size_kb": 0.0, "reason": str(exc),
                    "checked_at": datetime.now().isoformat(timespec="seconds"),
                    "duration_ms": 0.0,
                }
            results[idx] = res

            display = res["relative_path"]
            depth   = len(Path(display).parts) - 1
            indent  = "  " * depth + ("- " if depth else "")

            if res["status"] == STATUS_OK:
                colour, icon = GREEN, "[OK]"
            elif res["status"] in (STATUS_EMPTY, STATUS_NOT_PDF, STATUS_NO_PAGES):
                colour, icon = YELLOW, "[WARN]"
            else:
                colour, icon = RED, "[CORRUPT]"

            status_str = f"{colour}{icon:<10}{RESET}"
            print(
                f"{idx + 1:<5} {status_str} {res['pages']:>6}  "
                f"{res['size_kb']:>10.2f}  {res['duration_ms']:>6.0f}  {indent}{display}"
            )
            if res["reason"]:
                print(f"            {RED}  -> {res['reason']}{RESET}")

    return results


# ── Console summary ───────────────────────────────────────────────────────────
def print_summary(results: List[dict], root: Path) -> int:
    by_status = {s: [r for r in results if r["status"] == s]
                 for s in (STATUS_OK, STATUS_CORRUPT, STATUS_EMPTY, STATUS_NOT_PDF, STATUS_NO_PAGES)}

    print("\n" + "=" * 100)
    print(f"{BOLD}SUMMARY  —  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}{RESET}")
    print("=" * 100)
    print(f"  Root folder  : {root}")
    print(f"  Total PDFs   : {len(results)}")
    print(f"  {GREEN}OK       Valid    : {len(by_status[STATUS_OK])}{RESET}")
    print(f"  {RED}CORRUPT  Corrupt  : {len(by_status[STATUS_CORRUPT])}{RESET}")
    print(f"  {YELLOW}WARN     Empty    : {len(by_status[STATUS_EMPTY])}{RESET}")
    print(f"  {YELLOW}WARN     Not PDF  : {len(by_status[STATUS_NOT_PDF])}{RESET}")
    print(f"  {YELLOW}WARN     No pages : {len(by_status[STATUS_NO_PAGES])}{RESET}")

    for label, key in [("Corrupt", STATUS_CORRUPT), ("Empty", STATUS_EMPTY),
                        ("Invalid PDF header", STATUS_NOT_PDF), ("Zero-page", STATUS_NO_PAGES)]:
        group = by_status[key]
        if group:
            clr = RED if key == STATUS_CORRUPT else YELLOW
            print(f"\n{BOLD}{clr}{label} files:{RESET}")
            for r in group:
                print(f"  * {r['filepath']}")
                if r["reason"]:
                    print(f"    Reason: {r['reason']}")

    print("=" * 100)
    issues = sum(len(v) for k, v in by_status.items() if k != STATUS_OK)
    return 0 if issues == 0 else 1


# ── Excel export ──────────────────────────────────────────────────────────────
# Palette
_HEADER_FILL  = PatternFill("solid", fgColor="1F3864")   # dark navy
_OK_FILL      = PatternFill("solid", fgColor="E2EFDA")   # light green
_CORRUPT_FILL = PatternFill("solid", fgColor="FCE4D6")   # light red-orange
_WARN_FILL    = PatternFill("solid", fgColor="FFF2CC")   # light yellow
_ALT_FILL     = PatternFill("solid", fgColor="F5F5F5")   # zebra stripe
_SUMM_FILL    = PatternFill("solid", fgColor="D6E4F0")   # light blue summary
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

def _header_font(size=11):
    return Font(name="Arial", bold=True, color="FFFFFF", size=size)

def _body_font(bold=False):
    return Font(name="Arial", bold=bold, size=10)

def _status_fill(status: str) -> Optional[PatternFill]:
    if status == STATUS_OK:      return _OK_FILL
    if status == STATUS_CORRUPT: return _CORRUPT_FILL
    return _WARN_FILL   # EMPTY / NOT_A_PDF / NO_PAGES

def _center(cell):
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left(cell, wrap=False):
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=wrap)


def export_excel(results: List[dict], root: Path, out_path: Path) -> None:
    """Write a polished Excel report with a Details sheet and a Summary sheet."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: Details ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Details"
    ws.freeze_panes = "A3"

    # Title row
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = f"PDF Corruption Report  |  {root}  |  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    title_cell.font  = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    title_cell.fill  = PatternFill("solid", fgColor="1F3864")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    # Column headers
    headers = ["#", "Status", "Filename", "Folder", "Pages", "Size (KB)", "Duration (ms)", "Checked At", "Reason / Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font      = _header_font()
        cell.fill      = PatternFill("solid", fgColor="2E4057")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _BORDER
    ws.row_dimensions[2].height = 20

    # Data rows
    for i, r in enumerate(results, 1):
        row_num = i + 2
        status  = r["status"]
        row_fill = _status_fill(status) if status != STATUS_OK else (_ALT_FILL if i % 2 == 0 else None)

        values = [
            i,
            status,
            r["filename"],
            r["folder"],
            r["pages"],
            r["size_kb"],
            r["duration_ms"],
            r["checked_at"],
            r["reason"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font   = _body_font()
            cell.border = _BORDER
            if row_fill:
                cell.fill = row_fill
            if col in (1, 2, 5, 6, 7):
                _center(cell)
            else:
                _left(cell, wrap=(col == 9))

        # Bold the status cell
        ws.cell(row=row_num, column=2).font = _body_font(bold=True)

    # Column widths
    col_widths = [5, 12, 40, 40, 7, 11, 14, 20, 60]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Auto-filter
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{len(results) + 2}"

    # ── Sheet 2: Summary ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 14

    # Title
    ws2.merge_cells("A1:C1")
    t = ws2["A1"]
    t.value = "PDF Scan Summary"
    t.font  = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    t.fill  = PatternFill("solid", fgColor="1F3864")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    # Meta rows
    meta = [
        ("Root Folder",   str(root),                                      None),
        ("Scan Time",     datetime.now().strftime("%Y-%m-%d %H:%M:%S"),   None),
        ("Total PDFs",    len(results),                                    None),
    ]
    for ri, (label, value, _) in enumerate(meta, 2):
        lc = ws2.cell(row=ri, column=1, value=label)
        lc.font = _body_font(bold=True)
        lc.fill = _SUMM_FILL
        lc.border = _BORDER
        lc.alignment = Alignment(horizontal="left", vertical="center")

        vc = ws2.cell(row=ri, column=2, value=value)
        vc.font = _body_font()
        vc.fill = _SUMM_FILL
        vc.border = _BORDER
        vc.alignment = Alignment(horizontal="left", vertical="center")

        ws2.merge_cells(f"B{ri}:C{ri}")

    # Status breakdown header
    ws2.row_dimensions[6].height = 18
    for col, h in enumerate(["Status", "Count", "% of Total"], 1):
        c = ws2.cell(row=6, column=col, value=h)
        c.font  = _header_font(size=10)
        c.fill  = PatternFill("solid", fgColor="2E4057")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _BORDER

    status_rows = [
        (STATUS_OK,       "Valid",      "OK_FILL"),
        (STATUS_CORRUPT,  "Corrupt",    "CORRUPT_FILL"),
        (STATUS_EMPTY,    "Empty",      "WARN_FILL"),
        (STATUS_NOT_PDF,  "Not a PDF",  "WARN_FILL"),
        (STATUS_NO_PAGES, "No Pages",   "WARN_FILL"),
    ]
    total = len(results)
    for ri, (status, label, fill_name) in enumerate(status_rows, 7):
        count = sum(1 for r in results if r["status"] == status)
        pct   = count / total if total else 0
        row_fill = _status_fill(status)

        for col, val in enumerate([f"{status}  ({label})", count, pct], 1):
            c = ws2.cell(row=ri, column=col, value=val)
            c.font   = _body_font(bold=(col == 1))
            c.border = _BORDER
            c.fill   = row_fill or PatternFill("solid", fgColor="FFFFFF")
            if col == 3:
                c.number_format = "0.0%"
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 2:
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")

    # Totals row
    ri = 7 + len(status_rows)
    for col, val in enumerate(["TOTAL", f"=SUM(B7:B{ri - 1})", "100.0%"], 1):
        c = ws2.cell(row=ri, column=col, value=val)
        c.font   = _body_font(bold=True)
        c.border = _BORDER
        c.fill   = _SUMM_FILL
        c.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center")

    # ── Sheet 3: Issues only ──────────────────────────────────────────────────
    issues = [r for r in results if r["status"] != STATUS_OK]
    ws3 = wb.create_sheet("Issues Only")
    ws3.freeze_panes = "A3"

    ws3.merge_cells("A1:G1")
    t3 = ws3["A1"]
    t3.value = f"Issues Only — {len(issues)} file(s) need attention"
    t3.font  = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    t3.fill  = PatternFill("solid", fgColor="C00000")
    t3.alignment = Alignment(horizontal="left", vertical="center")
    ws3.row_dimensions[1].height = 24

    hdrs3 = ["#", "Status", "Filename", "Folder", "Size (KB)", "Checked At", "Reason"]
    for col, h in enumerate(hdrs3, 1):
        c = ws3.cell(row=2, column=col, value=h)
        c.font = _header_font()
        c.fill = PatternFill("solid", fgColor="2E4057")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _BORDER
    ws3.row_dimensions[2].height = 20

    for i, r in enumerate(issues, 1):
        row_num = i + 2
        vals = [i, r["status"], r["filename"], r["folder"], r["size_kb"], r["checked_at"], r["reason"]]
        fill = _status_fill(r["status"])
        for col, val in enumerate(vals, 1):
            c = ws3.cell(row=row_num, column=col, value=val)
            c.font   = _body_font(bold=(col == 2))
            c.border = _BORDER
            if fill:
                c.fill = fill
            if col in (1, 2, 5):
                _center(c)
            else:
                _left(c, wrap=(col == 7))

    for col, w in enumerate([5, 12, 40, 40, 11, 20, 60], 1):
        ws3.column_dimensions[get_column_letter(col)].width = w
    if issues:
        ws3.auto_filter.ref = f"A2:{get_column_letter(len(hdrs3))}{len(issues) + 2}"

    # Reorder sheets: Summary first
    wb.move_sheet("Summary", offset=-2)

    wb.save(out_path)
    print(f"\n{GREEN}{BOLD}Excel report saved:{RESET} {out_path}")


# ── CLI ───────────────────────────────────────────────────────────────────────
def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Recursively check PDF files for corruption and export an Excel report.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument("folder", nargs="?", help="Root folder to scan (prompted if omitted)")
    p.add_argument("--out", "-o", default="pdf_corruption_report.xlsx",
                   help="Output Excel file path (default: pdf_corruption_report.xlsx)")
    p.add_argument("--workers", "-w", type=int, default=4,
                   help="Parallel worker threads (default: 4)")
    p.add_argument("--no-excel", action="store_true", help="Skip Excel export")
    p.add_argument("--verbose", "-v", action="store_true", help="Enable debug logging")
    return p


def main() -> None:
    parser = build_parser()
    args   = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    folder = args.folder
    if not folder:
        folder = input("Enter the folder path to scan: ").strip().strip('"').strip("'")

    root = Path(folder).expanduser().resolve()

    if not root.exists():
        print(f"{RED}Error: Path does not exist → {root}{RESET}", file=sys.stderr)
        sys.exit(2)
    if not root.is_dir():
        print(f"{RED}Error: Path is not a directory → {root}{RESET}", file=sys.stderr)
        sys.exit(2)

    results   = scan_folder(root, workers=max(1, args.workers))
    exit_code = print_summary(results, root)

    if results and not args.no_excel:
        out_path = Path(args.out).expanduser().resolve()
        export_excel(results, root, out_path)

    sys.exit(exit_code)


if __name__ == "__main__":
    main()
