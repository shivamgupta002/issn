import os
import requests
import hashlib
from datetime import datetime
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook, Workbook
from tqdm import tqdm
from bs4 import BeautifulSoup

# ---------------- CONFIG ---------------- #
MAX_THREADS = 5
TIMEOUT = 15
CROSSREF_BASE = "https://api.crossref.org/works"
HEADERS = {
    "User-Agent": "DOI-Crawler/1.0 (mailto:your-email@example.com)"
}

# ---------------- SETUP ---------------- #
os.makedirs("downloads", exist_ok=True)
os.makedirs("output", exist_ok=True)

# ---------------- READ EXCEL WITH ERROR HANDLING ---------------- #
def read_input(file_path):
    """
    Reads DOI data from Excel file.
    Expected format: Column A = Source ID, Column B = DOI
    """
    if not os.path.exists(file_path):
        print(f"\n❌ ERROR: Input file not found: {file_path}")
        print(f"\n📁 Current directory: {os.getcwd()}")
        print(f"\n📋 Files in current directory:")
        for f in os.listdir("."):
            if f.endswith(('.xlsx', '.xls', '.csv')):
                print(f"  ✓ {f}")
        print(f"\n✏️  Solution:")
        print(f"  1. Create 'input_dois.xlsx' in: {os.getcwd()}")
        print(f"  2. Format: Column A = Source ID, Column B = DOI (starting from row 2)")
        raise FileNotFoundError(f"Input file '{file_path}' not found in {os.getcwd()}")

    try:
        wb = load_workbook(file_path)
        ws = wb.active
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            source_id, doi = row[0], row[1]
            if source_id and doi:
                data.append((str(source_id), str(doi).strip()))

        if not data:
            print(f"\n⚠️  WARNING: No data found in {file_path}")
            print(f"   Make sure data starts from Row 2 (Row 1 = headers)")
            return []

        print(f"✅ Loaded {len(data)} DOI entries from {file_path}")
        return data

    except Exception as e:
        print(f"\n❌ ERROR reading {file_path}: {e}")
        raise


# ---------------- FETCH METADATA VIA api.crossref.org/works/{doi} ---------------- #
def fetch_doi_metadata(doi):
    """
    Fetches full article metadata from:
        https://api.crossref.org/works/{doi}

    Returns the 'message' dict from CrossRef, or None on failure.
    """
    url = f"{CROSSREF_BASE}/{doi}"
    try:
        res = requests.get(url, headers=HEADERS, timeout=TIMEOUT)
        if res.status_code == 200:
            return res.json().get("message", None)
        elif res.status_code == 404:
            print(f"[NOT FOUND] CrossRef has no record for DOI: {doi}")
        else:
            print(f"[ERROR] CrossRef returned {res.status_code} for DOI: {doi}")
    except Exception as e:
        print(f"[ERROR] Failed to fetch CrossRef metadata for {doi}: {e}")
    return None


# ---------------- RESOLVE PDF ---------------- #
def resolve_pdf(doi, links):
    """Try to find a PDF URL from CrossRef link entries, then fallback to doi.org."""
    if links:
        for link in links:
            if link.get("content-type") == "application/pdf":
                return link.get("URL")

    try:
        headers = {"Accept": "application/pdf"}
        res = requests.get(
            f"https://doi.org/{doi}", headers=headers,
            allow_redirects=True, timeout=TIMEOUT
        )
        if "application/pdf" in res.headers.get("Content-Type", ""):
            return res.url
    except:
        pass

    return None


# -------- HELPER: GET META TAG WITH FALLBACK VARIATIONS -------- #
def get_meta_value(soup, primary_names, fallback_names=None):
    if fallback_names is None:
        fallback_names = []
    for name in primary_names + fallback_names:
        meta = soup.find("meta", attrs={"name": name})
        if meta and meta.get("content"):
            return meta.get("content", "")
    return ""


# -------- HELPER: GET MULTIPLE META TAG VALUES (FOR AUTHORS) -------- #
def get_meta_list(soup, tag_names):
    for tag_name in tag_names:
        metas = soup.find_all("meta", attrs={"name": tag_name})
        if metas:
            values = [m.get("content", "") for m in metas if m.get("content")]
            if values:
                return values
    return []


# -------- SCRAPE doi.org AS FALLBACK WHEN PDF NOT IN CROSSREF -------- #
def scrape_doi_metadata(doi):
    """
    Scrapes doi.org page for extra metadata when CrossRef doesn't have a PDF URL.
    Returns: authors, keywords, volume, issue, first_page, last_page, pdf_url
    """
    metadata = {
        "authors": "", "keywords": "", "volume": "", "issue": "",
        "first_page": "", "last_page": "", "pdf_url": ""
    }

    try:
        url = f"https://doi.org/{doi}"
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
        res = requests.get(url, headers=headers, timeout=TIMEOUT, allow_redirects=True)
        res.raise_for_status()
        soup = BeautifulSoup(res.content, "html.parser")

        # Authors
        authors_list = get_meta_list(soup, ["citation_author", "author", "DC.creator", "creator"])
        if not authors_list:
            authors_value = get_meta_value(soup, ["citation_authors"], ["authors", "DC.creator", "creator"])
            if authors_value:
                authors_list = [authors_value]
        metadata["authors"] = ", ".join(authors_list) if authors_list else ""

        # Keywords
        metadata["keywords"] = get_meta_value(
            soup, ["citation_keywords"], ["keywords", "DC.subject", "subject", "citation_subject"]
        )

        # Volume
        metadata["volume"] = get_meta_value(
            soup, ["citation_volume"], ["volume", "prism:volume", "journalvolume"]
        )

        # Issue
        metadata["issue"] = get_meta_value(
            soup, ["citation_issue"], ["issue", "prism:issueIdentifier", "journalissue"]
        )

        # Pages
        metadata["first_page"] = get_meta_value(
            soup, ["citation_firstpage"], ["citation_first_page", "startPage", "prism:startingPage"]
        )
        metadata["last_page"] = get_meta_value(
            soup, ["citation_lastpage"], ["citation_last_page", "endPage", "prism:endingPage"]
        )

        # PDF URL
        metadata["pdf_url"] = get_meta_value(
            soup, ["citation_pdf_url"], ["pdf_url", "pdfURL", "prism:url", "citation_pdf"]
        )

    except Exception as e:
        print(f"[WARNING] Failed to scrape doi.org for {doi}: {e}")

    return metadata


# ---------------- DOWNLOAD PDF ---------------- #
def download_pdf(url, source_id, doi, volume="", issue=""):
    """
    Downloads PDF into:  downloads/<source_id>/v{volume}_I{issue}/
    Filename:            <source_id>_<timestamp>.pdf
    """
    try:
        folder_name = f"v{volume}_I{issue}" if (volume or issue) else "no_volume_issue"
        folder_path = os.path.join("downloads", str(source_id), folder_name)
        os.makedirs(folder_path, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]
        filename = f"{source_id}_{timestamp}.pdf"
        path = os.path.join(folder_path, filename)

        res = requests.get(url, timeout=TIMEOUT)
        with open(path, "wb") as f:
            f.write(res.content)

        return os.path.join(str(source_id), folder_name, filename)

    except Exception as e:
        print(f"[ERROR] Failed to download PDF for {doi}: {e}")
        return None


# ---------------- PROCESS A SINGLE DOI ---------------- #
def process_doi(source_id, doi):
    """
    Full pipeline for one DOI:
      1. Fetch metadata from api.crossref.org/works/{doi}
      2. Extract all fields
      3. Try to resolve PDF (CrossRef links → doi.org header → doi.org scrape)
      4. Download PDF
      5. Return structured result
    """
    try:
        article = fetch_doi_metadata(doi)

        if article is None:
            return {
                "status": "Not Found in CrossRef", "error": "",
                "data": None, "doi": doi, "title": "", "source": "none"
            }

        # ---- Extract fields from CrossRef response ---- #
        title = article.get("title", [""])[0] if article.get("title") else ""

        authors = ", ".join([
            f"{a.get('family', '')} {a.get('given', '')}".strip()
            for a in article.get("author", [])
        ])

        published = article.get("issued", {}).get("date-parts", [[None]])[0]
        year  = published[0] if len(published) > 0 else ""
        month = published[1] if len(published) > 1 else ""
        day   = published[2] if len(published) > 2 else ""

        volume   = article.get("volume", "")
        issue    = article.get("issue", "")
        page_range = article.get("page", "")

        startpage, endpage = "", ""
        if page_range:
            if "-" in page_range:
                parts = page_range.split("-")
                startpage = parts[0]
                endpage   = parts[1] if len(parts) > 1 else ""
            else:
                startpage = endpage = page_range

        abstract    = article.get("abstract", "")
        ref_count   = article.get("reference-count", "")
        cited_by    = article.get("is-referenced-by-count", "")
        publisher   = article.get("publisher", "")
        journal     = article.get("container-title", [""])[0] if article.get("container-title") else ""
        issn_list   = ", ".join(article.get("ISSN", []))
        article_type = article.get("type", "")
        language    = article.get("language", "")
        subject     = ", ".join(article.get("subject", []))
        url         = article.get("URL", "")
        doi_actual  = article.get("DOI", doi)

        license_url = ""
        licenses = article.get("license", [])
        if licenses:
            license_url = licenses[0].get("URL", "")

        links = article.get("link", [])

        # ---- Resolve PDF ---- #
        pdf_url = resolve_pdf(doi_actual, links)
        doi_metadata = None

        if not pdf_url:
            print(f"[SCRAPING] No PDF in CrossRef, checking doi.org for {doi_actual}...")
            doi_metadata = scrape_doi_metadata(doi_actual)

            if doi_metadata["authors"] and not authors:
                authors = doi_metadata["authors"]
            if doi_metadata["volume"] and not volume:
                volume = doi_metadata["volume"]
            if doi_metadata["issue"] and not issue:
                issue = doi_metadata["issue"]
            if doi_metadata["first_page"] and not startpage:
                startpage = doi_metadata["first_page"]
            if doi_metadata["last_page"] and not endpage:
                endpage = doi_metadata["last_page"]
            if doi_metadata["pdf_url"]:
                pdf_url = doi_metadata["pdf_url"]

        if not pdf_url:
            return {
                "status": "PDF Not Found", "error": "",
                "data": None, "doi": doi_actual, "title": title, "source": "none"
            }

        # ---- Download PDF ---- #
        pdf_filename = download_pdf(pdf_url, source_id, doi_actual, volume, issue)

        if not pdf_filename:
            return {
                "status": "PDF Download Error", "error": "",
                "data": None, "doi": doi_actual, "title": title, "source": "none"
            }

        # Reconstruct page range if scraped
        if startpage and endpage and not page_range:
            page_range = f"{startpage}-{endpage}"
        elif startpage and not page_range:
            page_range = startpage

        return {
            "status": "Success",
            "error": "",
            "data": [
                source_id, title, doi_actual, url, pdf_url, pdf_filename,
                authors, year, month, day,
                journal, publisher, issn_list, article_type, language,
                volume, issue, page_range, startpage, endpage,
                abstract, ref_count, cited_by, license_url, subject
            ],
            "doi": doi_actual,
            "title": title,
            "source": "doi.org_scrape" if doi_metadata else "crossref"
        }

    except Exception as e:
        return {
            "status": "Error", "error": str(e),
            "data": None, "doi": doi, "title": "", "source": "none"
        }


# ============= REAL-TIME REPORT MANAGER ============= #
class ReportManager:
    """Creates and updates output Excel files and NotFound log in real-time."""

    def __init__(self):
        self.articles_file  = "output/articles_report.xlsx"
        self.status_file    = "output/status_report.xlsx"
        self.notfound_file  = "output/NotFound.txt"
        self._initialize_reports()

    def _initialize_reports(self):
        # Articles Report
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Source ID", "Title", "DOI", "CrossRef URL", "PDF URL", "PDF File",
            "Authors", "Year", "Month", "Day",
            "Journal", "Publisher", "ISSN List", "Article Type", "Language",
            "Volume", "Issue", "Page Range", "Start Page", "End Page",
            "Abstract", "Reference Count", "Cited By Count", "License URL", "Subjects"
        ])
        wb.save(self.articles_file)

        # Status Report
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.append(["Source ID", "DOI", "Title", "Status", "Error", "Source", "Timestamp"])
        wb2.save(self.status_file)

        # NotFound log
        with open(self.notfound_file, "w") as f:
            f.write("Source ID | DOI | Timestamp\n")
            f.write("=" * 80 + "\n")

    def append_article(self, article_data):
        try:
            wb = load_workbook(self.articles_file)
            ws = wb.active
            ws.append(article_data)
            wb.save(self.articles_file)
        except Exception as e:
            print(f"[ERROR] Failed to append article: {e}")

    def append_status(self, status_data):
        try:
            status_data.append(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            wb = load_workbook(self.status_file)
            ws = wb.active
            ws.append(status_data)
            wb.save(self.status_file)
        except Exception as e:
            print(f"[ERROR] Failed to append status: {e}")

    def append_notfound(self, source_id, doi):
        try:
            with open(self.notfound_file, "a") as f:
                f.write(f"{source_id} | {doi} | {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        except Exception as e:
            print(f"[ERROR] Failed to append to NotFound: {e}")


# ---------------- FINAL SUMMARY ---------------- #
def save_outputs(articles_count, status_count, not_found_count):
    print("\n" + "=" * 60)
    print("✅ FINAL REPORT SUMMARY")
    print("=" * 60)
    print(f"📄 Articles Report : output/articles_report.xlsx  ({articles_count} rows)")
    print(f"📄 Status Report   : output/status_report.xlsx    ({status_count} rows)")
    print(f"📄 Not Found List  : output/NotFound.txt          ({not_found_count} items)")
    print(f"📁 PDFs Saved      : downloads/<SourceID>/v<volume>_I<issue>/")
    print("=" * 60 + "\n")


# ---------------- MAIN RUNNER ---------------- #
def run_crawler(input_file):
    input_data = read_input(input_file)
    if not input_data:
        return

    report_manager = ReportManager()
    print(f"\n📊 Real-time report generation initialized")
    print(f"   → {len(input_data)} DOIs to process\n")

    articles_count  = 0
    status_count    = 0
    not_found_count = 0

    with ThreadPoolExecutor(max_workers=MAX_THREADS) as executor:
        futures = {
            executor.submit(process_doi, source_id, doi): (source_id, doi)
            for source_id, doi in input_data
        }

        for future in tqdm(as_completed(futures), total=len(futures), desc="Processing DOIs"):
            source_id, doi = futures[future]
            result = future.result()

            # ---- Append status row ---- #
            status_row = [
                source_id, result["doi"], result["title"],
                result["status"], result["error"], result["source"]
            ]
            report_manager.append_status(status_row)
            status_count += 1

            # ---- Append article if success ---- #
            if result["data"]:
                report_manager.append_article(result["data"])
                articles_count += 1
                print(f"  ✅ [{source_id}] {result['title'][:60]}...")

            else:
                if result["status"] in ("PDF Not Found", "Not Found in CrossRef"):
                    report_manager.append_notfound(source_id, result["doi"])
                    not_found_count += 1
                print(f"  ⚠️  [{source_id}] {result['status']} — {result['doi']}")

    save_outputs(articles_count, status_count, not_found_count)


# ---------------- ENTRY POINT ---------------- #
if __name__ == "__main__":
    import sys

    if len(sys.argv) > 1:
        input_file = sys.argv[1]
    else:
        default_files = ["input_dois.xlsx", "dois.xlsx", "input.xlsx", "data.xlsx"]
        existing_files = [f for f in default_files if os.path.exists(f)]

        print("\n" + "=" * 60)
        print("DOI Article Crawler  (api.crossref.org/works/{doi})")
        print("=" * 60)

        if existing_files:
            print(f"\n✓ Found {len(existing_files)} Excel file(s):")
            for i, f in enumerate(existing_files, 1):
                print(f"  {i}. {f}")

            if len(existing_files) == 1:
                input_file = existing_files[0]
                print(f"\n→ Using: {input_file}")
            else:
                choice = input(f"\nSelect file (1-{len(existing_files)}): ").strip()
                try:
                    input_file = existing_files[int(choice) - 1]
                except:
                    input_file = existing_files[0]
        else:
            print(f"\n❌ No Excel files found in: {os.getcwd()}")
            print(f"\n📝 Create 'input_dois.xlsx' with:")
            print(f"   Row 1 (Headers): Source ID | DOI")
            print(f"   Row 2 onwards:   1 | 10.1038/s41586-021-03819-2")
            print(f"\n💡 Or run with: python doi_crawler.py your_file.xlsx")
            sys.exit(1)

    try:
        run_crawler(input_file)
    except KeyboardInterrupt:
        print("\n\n⚠️  Process interrupted by user")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ Fatal error: {e}")
        sys.exit(1)
