import os
import requests
import hashlib
from datetime import datetime
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook, Workbook
from tqdm import tqdm
from bs4 import BeautifulSoup

# ---------------- CONFIG ---------------- #
MAX_THREADS = 5
TIMEOUT = 15

# ---------------- SETUP ---------------- #
os.makedirs("downloads", exist_ok=True)
os.makedirs("output", exist_ok=True)

# ---------------- READ EXCEL WITH ERROR HANDLING ---------------- #
def read_input(file_path):
    """
    Reads ISSN data from Excel file with proper error handling.
    Expected format: Column A = Source ID, Column B = ISSN
    """
    # Check if file exists
    if not os.path.exists(file_path):
        print(f"\n❌ ERROR: Input file not found: {file_path}")
        print(f"\n📁 Current directory: {os.getcwd()}")
        print(f"\n📋 Files in current directory:")
        files = os.listdir(".")
        for f in files:
            if f.endswith(('.xlsx', '.xls', '.csv')):
                print(f"   ✓ {f}")
        
        print(f"\n✏️  Solution:")
        print(f"   1. Create 'input.xlsx' in: {os.getcwd()}")
        print(f"   2. OR update the filename in the script at the bottom")
        print(f"   3. Format: Column A = Source ID, Column B = ISSN (starting from row 2)")
        raise FileNotFoundError(f"Input file '{file_path}' not found in {os.getcwd()}")
    
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        data = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            source_id, issn = row
            if source_id and issn:
                data.append((str(source_id), str(issn)))

        if not data:
            print(f"\n⚠️  WARNING: No data found in {file_path}")
            print(f"   Make sure data starts from Row 2 (Row 1 should be headers)")
            return []

        print(f"✅ Loaded {len(data)} ISSN entries from {file_path}")
        return data

    except Exception as e:
        print(f"\n❌ ERROR reading {file_path}: {e}")
        raise

# ---------------- FETCH ARTICLES (UPDATED LOGIC) ---------------- #
def fetch_articles(issn):
    articles = []
    cursor = "*"

    headers = {
        "User-Agent": "ISSN-Crawler/1.0 (mailto:your-email@example.com)"
    }

    while True:
        url = f"https://api.crossref.org/journals/{issn}/works"
        params = {
            "rows": 1000,
            "cursor": cursor,
            "sort": "published",
            "order": "desc"
        }

        try:
            res = requests.get(url, params=params, headers=headers, timeout=TIMEOUT)

            if res.status_code != 200:
                print(f"[ERROR] API failed: {res.status_code} for ISSN {issn}")
                break

            data = res.json()
            items = data["message"]["items"]

            if not items:
                break

            articles.extend(items)
            cursor = data["message"].get("next-cursor")

        except Exception as e:
            print(f"[ERROR] API fetch failed for ISSN {issn}: {e}")
            break

    if not articles:
        return []

    # --------- FILTER LAST 2 YEARS BASED ON AVAILABLE DATA --------- #
    years = []

    for art in articles:
        pub = art.get("issued", {}).get("date-parts", [[None]])[0]
        if pub and pub[0]:
            years.append(pub[0])

    if not years:
        return []

    latest_year = max(years)
    second_latest_year = latest_year - 1

    print(f"Latest years for ISSN {issn}: {latest_year}, {second_latest_year}")

    filtered_articles = []

    for art in articles:
        pub = art.get("issued", {}).get("date-parts", [[None]])[0]
        if pub and pub[0] in [latest_year, second_latest_year]:
            filtered_articles.append(art)

    print(f"Filtered {len(filtered_articles)} articles for last 2 available years")

    return filtered_articles

# ---------------- RESOLVE PDF ---------------- #
def resolve_pdf(doi, links):
    if links:
        for link in links:
            if link.get("content-type") == "application/pdf":
                return link.get("URL")

    try:
        headers = {"Accept": "application/pdf"}
        res = requests.get(f"https://doi.org/{doi}", headers=headers, allow_redirects=True, timeout=TIMEOUT)

        if "application/pdf" in res.headers.get("Content-Type", ""):
            return res.url
    except:
        pass

    return None

# -------- HELPER: GET META TAG WITH FALLBACK VARIATIONS -------- #
def get_meta_value(soup, primary_names, fallback_names=None):
    """
    Get meta tag value with intelligent fallback.
    
    Args:
        soup: BeautifulSoup object
        primary_names: List of primary tag names to try
        fallback_names: List of fallback tag names if primary not found
    
    Returns:
        String value or empty string if not found
    """
    if fallback_names is None:
        fallback_names = []
    
    all_names = primary_names + fallback_names
    
    # Try primary names first
    for name in primary_names:
        meta = soup.find("meta", attrs={"name": name})
        if meta and meta.get("content"):
            return meta.get("content", "")
    
    # Try fallback names
    for name in fallback_names:
        meta = soup.find("meta", attrs={"name": name})
        if meta and meta.get("content"):
            return meta.get("content", "")
    
    return ""

# -------- HELPER: GET MULTIPLE META TAG VALUES (FOR AUTHORS) -------- #
def get_meta_list(soup, tag_names):
    """
    Get multiple meta tag values (for fields that can have multiple entries like authors).
    
    Args:
        soup: BeautifulSoup object
        tag_names: List of tag names to search for (in order of preference)
    
    Returns:
        List of values found
    """
    for tag_name in tag_names:
        metas = soup.find_all("meta", attrs={"name": tag_name})
        if metas:
            values = [m.get("content", "") for m in metas if m.get("content")]
            if values:
                return values
    
    return []

# -------- MAIN SCRAPING FUNCTION WITH FALLBACK VARIATIONS -------- #
def scrape_doi_metadata(doi):
    """
    Scrapes metadata from doi.org page when CrossRef PDF lookup fails.
    Extracts: authors, keywords, volume, issue, first page, last page, PDF URL
    
    Uses intelligent fallback to find similar tags if primary ones don't exist.
    """
    metadata = {
        "authors": "",
        "keywords": "",
        "volume": "",
        "issue": "",
        "first_page": "",
        "last_page": "",
        "pdf_url": ""
    }

    try:
        url = f"https://doi.org/{doi}"
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        }
        
        res = requests.get(url, headers=headers, timeout=TIMEOUT, allow_redirects=True)
        res.raise_for_status()
        
        soup = BeautifulSoup(res.content, "html.parser")

        # ===== AUTHORS (Multiple variations) =====
        # Primary: citation_author (multiple tags)
        # Fallbacks: citation_authors, author, creator, DC.creator
        authors_list = get_meta_list(soup, [
            "citation_author",      # Standard multiple author tags
            "author",               # Generic author
            "DC.creator",           # Dublin Core standard
            "creator",              # Generic creator
        ])
        
        if not authors_list:
            # Try single-value fallbacks
            authors_value = get_meta_value(soup,
                ["citation_authors"],  # Single combined authors
                ["authors", "DC.creator", "creator"]
            )
            if authors_value:
                authors_list = [authors_value]
        
        metadata["authors"] = ", ".join(authors_list) if authors_list else ""

        # ===== KEYWORDS (Multiple variations) =====
        # Primary: citation_keywords
        # Fallbacks: keywords, DC.subject, subject
        metadata["keywords"] = get_meta_value(soup,
            ["citation_keywords"],
            ["keywords", "DC.subject", "subject", "citation_subject"]
        )

        # ===== VOLUME (Multiple variations) =====
        # Primary: citation_volume
        # Fallbacks: volume, prism:volume, journalvolume
        metadata["volume"] = get_meta_value(soup,
            ["citation_volume"],
            ["volume", "prism:volume", "journalvolume"]
        )

        # ===== ISSUE (Multiple variations) =====
        # Primary: citation_issue
        # Fallbacks: issue, prism:issueIdentifier, journalissue
        metadata["issue"] = get_meta_value(soup,
            ["citation_issue"],
            ["issue", "prism:issueIdentifier", "journalissue", "citation_issue_identifier"]
        )

        # ===== FIRST PAGE (Multiple variations) =====
        # Primary: citation_firstpage
        # Fallbacks: citation_first_page, startPage, prism:startingPage
        metadata["first_page"] = get_meta_value(soup,
            ["citation_firstpage"],
            ["citation_first_page", "startPage", "prism:startingPage", "pageStart"]
        )

        # ===== LAST PAGE (Multiple variations) =====
        # Primary: citation_lastpage
        # Fallbacks: citation_last_page, endPage, prism:endingPage
        metadata["last_page"] = get_meta_value(soup,
            ["citation_lastpage"],
            ["citation_last_page", "endPage", "prism:endingPage", "pageEnd"]
        )

        # ===== PDF URL (Multiple variations) =====
        # Primary: citation_pdf_url
        # Fallbacks: pdf_url, pdfURL, prism:url, citation_pdf
        metadata["pdf_url"] = get_meta_value(soup,
            ["citation_pdf_url"],
            ["pdf_url", "pdfURL", "prism:url", "citation_pdf", "citation_fulltext_pdf_url"]
        )

        return metadata

    except Exception as e:
        print(f"[WARNING] Failed to scrape doi.org for {doi}: {e}")
        return metadata

# ---------------- DOWNLOAD PDF WITH NEW FOLDER STRUCTURE ---------------- #
def download_pdf(url, source_id, doi, volume="", issue=""):
    """
    Download PDF and save to folder structure: downloads/SourceID/v{volume}_I{issue}/
    Filename format: sourceID_timestamp.pdf
    """
    try:
        # Create folder structure
        folder_name = f"v{volume}_I{issue}" if volume or issue else "no_volume_issue"
        folder_path = os.path.join("downloads", str(source_id), folder_name)
        os.makedirs(folder_path, exist_ok=True)

        # Create filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]  # YYYYmmdd_HHMMSS_ms
        filename = f"{source_id}_{timestamp}.pdf"
        path = os.path.join(folder_path, filename)

        # Download PDF
        res = requests.get(url, timeout=TIMEOUT)
        with open(path, "wb") as f:
            f.write(res.content)

        # Return relative path for Excel
        return os.path.join(str(source_id), folder_name, filename)
    except Exception as e:
        print(f"[ERROR] Failed to download PDF for {doi}: {e}")
        return None

# ---------------- PROCESS ARTICLE (ENHANCED) ---------------- #
def process_article(article, source_id, issn):
    try:
        title = article.get("title", [""])[0]
        doi = article.get("DOI", "")
        
        # Get authors from CrossRef first
        authors = ", ".join([
            f"{a.get('family', '')} {a.get('given', '')}".strip()
            for a in article.get("author", [])
        ])

        published = article.get("issued", {}).get("date-parts", [[None]])[0]
        year = published[0] if len(published) > 0 else ""
        month = published[1] if len(published) > 1 else ""
        day = published[2] if len(published) > 2 else ""

        volume = article.get("volume", "")
        issue = article.get("issue", "")

        pageRange = article.get("page", "")
        startpage = ""
        endpage = ""
        if pageRange:
            if "-" in pageRange:
                parts = pageRange.split("-")
                startpage = parts[0]
                endpage = parts[1] if len(parts) > 1 else ""
            else:
                startpage = pageRange
                endpage = pageRange

        abstract = article.get("abstract", "")
        ref_count = article.get("reference-count", "")
        cited_by = article.get("is-referenced-by-count", "")
        publisher = article.get("publisher", "")
        journal = article.get("container-title", [""])[0]
        issn_list = ", ".join(article.get("ISSN", []))
        article_type = article.get("type", "")
        language = article.get("language", "")
        license_url = ""
        licenses = article.get("license", [])
        if licenses:
            license_url = licenses[0].get("URL", "")
        subject = ", ".join(article.get("subject", []))
        url = article.get("URL", "")

        links = article.get("link", [])
        pdf_url = resolve_pdf(doi, links)

        # ========== NEW: FALLBACK TO DOI.ORG SCRAPING ========== #
        doi_metadata = None
        if not pdf_url:
            print(f"[SCRAPING] No PDF in CrossRef, checking doi.org for {doi}...")
            doi_metadata = scrape_doi_metadata(doi)
            
            # Use scraped data if available
            if doi_metadata["authors"] and not authors:
                authors = doi_metadata["authors"]
            if doi_metadata["volume"] and not volume:
                volume = doi_metadata["volume"]
            if doi_metadata["issue"] and not issue:
                issue = doi_metadata["issue"]
            if doi_metadata["first_page"] and not startpage:
                startpage = doi_metadata["first_page"]
            if doi_metadata["last_page"] and not endpage:
                endpage = doi_metadata["last_page"]
            if doi_metadata["pdf_url"]:
                pdf_url = doi_metadata["pdf_url"]

        if not pdf_url:
            return {
                "status": "PDF Not Found", "error": "",
                "data": None, "doi": doi, "title": title, "source": "none"
            }

        pdf_filename = download_pdf(pdf_url, source_id, doi, volume, issue)

        if not pdf_filename:
            return {
                "status": "PDF Download Error", "error": "",
                "data": None, "doi": doi, "title": title, "source": "none"
            }

        # Reconstruct pageRange if we have start/end
        if startpage and endpage and not pageRange:
            pageRange = f"{startpage}-{endpage}"
        elif startpage and not pageRange:
            pageRange = startpage

        return {
            "status": "Success",
            "error": "",
            "data": [
                source_id, title, doi, url, pdf_url, pdf_filename,
                authors, year, month, day,
                journal, publisher, issn_list, article_type, language,
                volume, issue, pageRange, startpage, endpage,
                abstract, ref_count, cited_by, license_url, subject
            ],
            "doi": doi,
            "title": title,
            "source": "doi.org" if doi_metadata else "crossref"
        }

    except Exception as e:
        return {
            "status": "Error", "error": str(e),
            "data": None, "doi": "", "title": "", "source": "none"
        }

# ============= NEW: REAL-TIME ISSN STATUS UPDATES ============= #
class ISSNStatusTracker:
    """Tracks and updates ISSN processing status in real-time"""
    
    def __init__(self, output_file="output/issn_status_realtime.xlsx"):
        self.output_file = output_file
        self.issn_statuses = {}
        self.lock = None
        
    def initialize(self, issn_data):
        """Initialize Excel with all ISSNs and starting status"""
        wb = Workbook()
        ws = wb.active
        ws.title = "ISSN Status"
        ws.append(["Source ID", "ISSN", "Status", "Articles Found", "PDFs Downloaded", "Errors", "Last Updated"])
        
        for source_id, issn in issn_data:
            ws.append([source_id, issn, "Pending", 0, 0, 0, ""])
            self.issn_statuses[issn] = {
                "source_id": source_id,
                "status": "Pending",
                "articles": 0,
                "pdfs": 0,
                "errors": 0
            }
        
        wb.save(self.output_file)
        print(f"\n📊 ISSN Status tracker initialized: {self.output_file}")
    
    def update_issn_status(self, source_id, issn, status, articles_count=0, pdfs_count=0, errors_count=0):
        """Update ISSN status in real-time Excel"""
        try:
            wb = load_workbook(self.output_file)
            ws = wb.active
            
            # Find and update the row
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if row[1].value == issn:  # Match ISSN
                    row[2].value = status  # Status
                    row[3].value = articles_count  # Articles
                    row[4].value = pdfs_count  # PDFs
                    row[5].value = errors_count  # Errors
                    row[6].value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # Timestamp
                    break
            
            wb.save(self.output_file)
            
        except Exception as e:
            print(f"[WARNING] Could not update status for {issn}: {e}")

# ---------------- MAIN RUN WITH REAL-TIME UPDATES & REPORTS ------------ #
def run_crawler(input_file):
    input_data = read_input(input_file)
    
    # Initialize trackers and report manager
    status_tracker = ISSNStatusTracker()
    status_tracker.initialize(input_data)
    
    report_manager = ReportManager()
    print("\n📊 Real-time report generation initialized")

    articles_count = 0
    status_count = 0
    not_found_count = 0

    for source_id, issn in input_data:
        print(f"\n{'='*60}")
        print(f"Processing ISSN: {issn} (Source ID: {source_id})")
        print(f"{'='*60}")

        articles = fetch_articles(issn)

        if not articles:
            print(f"[WARNING] No articles found for ISSN {issn}")
            status_tracker.update_issn_status(source_id, issn, "No Articles Found", 0, 0, 0)
            continue

        articles_found = len(articles)
        pdfs_downloaded = 0
        errors_count = 0

        with ThreadPoolExecutor(max_workers=MAX_THREADS) as executor:
            futures = [
                executor.submit(process_article, art, source_id, issn)
                for art in articles
            ]

            for future in tqdm(as_completed(futures), total=len(futures), desc=f"Processing {issn}"):
                result = future.result()

                # ========== REAL-TIME: Append status immediately ========== #
                status_row = [
                    source_id, issn, result["doi"], result["title"],
                    result["status"], result["error"]
                ]
                report_manager.append_status(status_row)
                status_count += 1

                # ========== REAL-TIME: Append article if successful ========== #
                if result["data"]:
                    report_manager.append_article(result["data"])
                    articles_count += 1
                    pdfs_downloaded += 1
                else:
                    # ========== REAL-TIME: Append to NotFound if PDF not found ========== #
                    if result["status"] == "PDF Not Found":
                        report_manager.append_notfound(source_id, result['doi'])
                        not_found_count += 1
                    errors_count += 1

        # Update ISSN status after completion
        status_tracker.update_issn_status(
            source_id, issn, "Completed",
            articles_found, pdfs_downloaded, errors_count
        )
        
        print(f"\n✅ ISSN {issn} completed: {pdfs_downloaded}/{articles_found} PDFs downloaded")
        print(f"   📊 Running totals - Articles: {articles_count} | Status: {status_count} | Not Found: {not_found_count}")

    # Final summary
    save_outputs(articles_count, status_count, not_found_count)

# ============= NEW: REAL-TIME REPORT GENERATION ============= #
class ReportManager:
    """Manages real-time report generation and updates"""
    
    def __init__(self):
        self.articles_file = "output/articles_report.xlsx"
        self.status_file = "output/status_report.xlsx"
        self.notfound_file = "output/NotFound.txt"
        self._initialize_reports()
    
    def _initialize_reports(self):
        """Create empty Excel files with headers"""
        # Articles Report
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Source ID", "Title", "DOI", "CrossRef URL", "PDF URL", "PDF File",
            "Authors", "Year", "Month", "Day",
            "Journal", "Publisher", "ISSN List", "Article Type", "Language",
            "Volume", "Issue", "Page Range", "Start Page", "End Page",
            "Abstract", "Reference Count", "Cited By Count", "License URL", "Subjects"
        ])
        wb.save(self.articles_file)
        
        # Status Report
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.append(["SourceID", "ISSN", "DOI", "Title", "Status", "Error", "Timestamp"])
        wb2.save(self.status_file)
        
        # NotFound Text File
        with open(self.notfound_file, "w") as f:
            f.write("Source ID | DOI | Timestamp\n")
            f.write("="*80 + "\n")
    
    def append_article(self, article_data):
        """Append a single article row to articles_report.xlsx"""
        try:
            wb = load_workbook(self.articles_file)
            ws = wb.active
            ws.append(article_data)
            wb.save(self.articles_file)
        except Exception as e:
            print(f"[ERROR] Failed to append article: {e}")
    
    def append_status(self, status_data):
        """Append a single status row to status_report.xlsx"""
        try:
            status_data.append(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            wb = load_workbook(self.status_file)
            ws = wb.active
            ws.append(status_data)
            wb.save(self.status_file)
        except Exception as e:
            print(f"[ERROR] Failed to append status: {e}")
    
    def append_notfound(self, source_id, doi):
        """Append to NotFound.txt file"""
        try:
            with open(self.notfound_file, "a") as f:
                f.write(f"{source_id} | {doi} | {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        except Exception as e:
            print(f"[ERROR] Failed to append to NotFound: {e}")

# ---------------- SAVE OUTPUT (FINAL SUMMARY) ------------ #
def save_outputs(articles, status, not_found):
    """Final summary - reports already saved in real-time"""
    print("\n" + "="*60)
    print("✅ FINAL REPORT SUMMARY")
    print("="*60)
    print(f"📄 Articles Report: output/articles_report.xlsx ({len(articles)} rows)")
    print(f"📄 Status Report: output/status_report.xlsx ({len(status)} rows)")
    print(f"📄 Not Found List: output/NotFound.txt ({len(not_found)} items)")
    print(f"📁 PDFs Saved: downloads/<SourceID>/v<volume>_I<issue>/")
    print("="*60 + "\n")

# ---------------- ENTRY WITH HELP ---------------- #
if __name__ == "__main__":
    import sys
    
    # Allow command line argument for input file
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
    else:
        # Try default filenames
        default_files = ["input.xlsx", "data.xlsx", "issn.xlsx", "articles.xlsx"]
        input_file = None
        
        print("\n" + "="*60)
        print("ISSN Article Crawler")
        print("="*60)
        
        # Check for existing files
        existing_files = [f for f in default_files if os.path.exists(f)]
        
        if existing_files:
            print(f"\n✓ Found {len(existing_files)} Excel file(s):")
            for i, f in enumerate(existing_files, 1):
                print(f"  {i}. {f}")
            
            if len(existing_files) == 1:
                input_file = existing_files[0]
                print(f"\n→ Using: {input_file}")
            else:
                choice = input(f"\nSelect file (1-{len(existing_files)}): ").strip()
                try:
                    input_file = existing_files[int(choice) - 1]
                except:
                    input_file = existing_files[0]
        else:
            print(f"\n❌ No Excel files found in current directory")
            print(f"\n📍 Expected location: {os.getcwd()}")
            print(f"\n📝 Create a file named 'input.xlsx' with:")
            print(f"   Row 1 (Headers):  Source ID  |  ISSN")
            print(f"   Row 2 onwards:    1          |  0028-0836")
            print(f"                     2          |  2051-1426")
            print(f"\n💡 Or run with: python script.py your_file.xlsx")
            sys.exit(1)
    
    try:
        run_crawler(input_file)
    except KeyboardInterrupt:
        print("\n\n⚠️  Process interrupted by user")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ Fatal error: {e}")
        sys.exit(1)
